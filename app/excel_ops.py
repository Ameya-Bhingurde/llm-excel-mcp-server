from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook


def load_excel(path: str | Path, sheet_name: str) -> pd.DataFrame:
    """
    Load an Excel worksheet into a pandas DataFrame.
    """
    path = Path(path)
    if not path.exists():
        raise ValueError(f"Excel file not found at '{path}'.")

    try:
        df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError as exc:
        raise ValueError(f"Worksheet '{sheet_name}' not found in '{path.name}'.") from exc
    except Exception as exc:
        raise ValueError(f"Failed to load Excel file: {exc}") from exc
    return df


def clean_sheet(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Clean a DataFrame and return the cleaned DF plus a summary of changes.
    """
    original_rows = len(df)
    summary = {
        "original_rows": original_rows,
        "rows_removed": 0,
        "columns_normalized": [],
        "nulls_filled": 0  # We generally drop rows, but if we filled via mean later we'd track it
    }

    if df.empty:
        summary["final_rows"] = 0
        return df, summary

    # 1. Normalise column names
    old_cols = list(df.columns)
    df = df.rename(columns=lambda c: c.strip() if isinstance(c, str) else c)
    new_cols = list(df.columns)
    
    # Simple check for changes
    if old_cols != new_cols:
        summary["columns_normalized"] = [o for o, n in zip(old_cols, new_cols) if o != n]

    # 2. Drop rows that are entirely NaN
    cleaned = df.dropna(how="all")
    summary["rows_removed"] = original_rows - len(cleaned)

    # 3. Infer types
    cleaned = cleaned.convert_dtypes()
    
    summary["final_rows"] = len(cleaned)
    return cleaned, summary


def profile_data(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Produce a light-weight profile of the DataFrame.
    """
    # Simply identify numeric vs categorical for charting suggestions
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    categorical_cols = df.select_dtypes(exclude=['number']).columns.tolist()

    return {
        "row_count": int(df.shape[0]),
        "column_count": int(df.shape[1]),
        "columns": list(df.columns.astype(str)),
        "numeric_columns": [str(c) for c in numeric_cols],
        "categorical_columns": [str(c) for c in categorical_cols],
        "null_counts": {str(col): int(df[col].isna().sum()) for col in df.columns},
    }


def prepare_chart_data(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Prepare data for visual analysis in the UI.
    If data is too large, it suggests aggregation or sampling.
    """
    # For now, just return the profile + a limited preview for charts if very large
    # Streamlit can handle decent size, but let's be safe.
    
    limit = 1000
    is_truncated = len(df) > limit
    chart_df = df.head(limit) if is_truncated else df
    
    # We return the records so the frontend can create a new dataframe or rendering
    return {
        "is_truncated": is_truncated,
        "limit": limit,
        "data": chart_df.to_dict(orient="records"),
        "profile": profile_data(df)
    }


def create_pivot_table(
    df: pd.DataFrame,
    index: List[str],
    values: List[str],
    aggfunc: str = "sum",
) -> pd.DataFrame:
    """
    Create a pivot table, handling case-insensitive column matching.
    """
    # Map input columns to actual DataFrame columns (case-insensitive)
    def match_cols(requested: List[str], actual: pd.Index) -> List[str]:
        matched = []
        actual_map = {c.lower(): c for c in actual}
        for req in requested:
            if req.lower() in actual_map:
                matched.append(actual_map[req.lower()])
            else:
                # If not found, let pandas raise the error or raise our own
                # But we'll pass original to let pandas complain if it's truly missing
                matched.append(req) 
        return matched

    real_index = match_cols(index, df.columns)
    real_values = match_cols(values, df.columns)
    
    # Normalize aggfunc
    valid_aggs = {"sum", "mean", "count", "min", "max", "average"} # average -> mean
    agg_lower = aggfunc.lower()
    if agg_lower == "average":
        agg_lower = "mean"
        
    if agg_lower not in valid_aggs:
        raise ValueError(f"Unsupported aggfunc '{aggfunc}'. Supported: {sorted(valid_aggs)}")

    missing_index = [c for c in real_index if c not in df.columns]
    missing_values = [c for c in real_values if c not in df.columns]
    
    if missing_index or missing_values:
         raise ValueError(
            f"Missing columns: index={missing_index}, values={missing_values}. "
            f"Available: {list(df.columns)}"
        )

    pivot = pd.pivot_table(
        df,
        index=real_index,
        values=real_values,
        aggfunc=agg_lower,
    ).reset_index()
    
    return pivot


def insert_formula(path: str | Path, sheet: str, cell: str, formula: str) -> None:
    """
    Insert a formula into a specific cell.
    """
    path = Path(path)
    if not path.exists():
        raise ValueError(f"Excel file not found at '{path}'.")

    try:
        wb = load_workbook(path)
    except Exception as exc:
        raise ValueError(f"Failed to open workbook: {exc}") from exc

    if sheet not in wb.sheetnames:
        raise ValueError(f"Worksheet '{sheet}' not found in '{path.name}'.")

    ws = wb[sheet]
    ws[cell] = formula
    wb.save(path)


def save_excel(df: pd.DataFrame, path: str | Path, sheet_name: str = "Sheet1") -> None:
    """
    Save a DataFrame back to an Excel file.
    """
    path = Path(path)
    try:
        # Use existing file if possible to avoid overwriting other sheets?
        # For this project, we might overwrite. simpler.
        # But 'clean_excel' usually replaces the sheet. 
        # Let's stick to full overwrite for safety/simplicity as per original code context,
        # unless we want to be fancy. The original code did overwrite.
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as exc:
        raise ValueError(f"Failed to save Excel file: {exc}") from exc
